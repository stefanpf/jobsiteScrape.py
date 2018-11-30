#! python3

# jobsiteScrape.py - scrapes a number of German job websites for
# placements,  selects the interesting ones based on some hard-coded
# criteria, and saves  the results in an easy to read spreadsheet.
#
# v0.6 - 2018-11-30

import os
import datetime
import time
import re
import openpyxl
from openpyxl.styles import Alignment, Font
from scrapers import scrape_stepstone, scrape_bsj

# Results should come as a nested dictionary like so:
# {<jobTitle> : {
#       <company> : value,
#       <desc> : value,
#       <link> : value,
#       <date> : value}}


def clean_results(results):

    multi_newlines_regex = re.compile(r'\n{3,}')

    for key, value in results.items():
        value['desc'] = re.sub(multi_newlines_regex, "\n\n", value['desc'])
    print('Cleaning up results...')

    return results


def write_to_XLS(data_dictionary):

    try:
        wb = openpyxl.load_workbook('Results %s.xlsx'
                                    % datetime.datetime.today()
                                    .strftime('%Y-%m-%d'))
        sheet = wb.active
        print("Opened spreadsheet to write results.")

    except FileNotFoundError:
        print("Spreadsheet not found, creating a new spreadsheet...")
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 120
        sheet.column_dimensions['E'].width = 30
        top_row_style_obj = Font(name='Arial', bold=True, size=12)
        sheet['A1'].font = top_row_style_obj
        sheet['B1'].font = top_row_style_obj
        sheet['C1'].font = top_row_style_obj
        sheet['D1'].font = top_row_style_obj
        sheet['E1'].font = top_row_style_obj
        sheet['A1'].value = "Job Title"
        sheet['B1'].value = "Company"
        sheet['C1'].value = "Date"
        sheet['D1'].value = "Description"
        sheet['E1'].value = "Link"
        wb.save('Results %s.xlsx'
                % datetime.datetime.today().strftime('%Y-%m-%d'))
        print("New spreadsheet created.")
        wb = openpyxl.load_workbook('Results %s.xlsx'
                                    % datetime.datetime.today()
                                    .strftime('%Y-%m-%d'))
        sheet = wb.active
        print("Opened spreadsheet to write results.")

    row_num = 2
    for row in range(1, sheet.max_row):
        if sheet.cell(row=row, column=1):
            row_num += 1

    for key, value in data_dictionary.items():
        sheet.cell(row=row_num, column=1).value = key
        sheet.cell(row=row_num, column=1).alignment = Alignment(vertical='top')
        sheet.cell(row=row_num, column=2).value = value['company']
        sheet.cell(row=row_num, column=2).alignment = Alignment(vertical='top')
        sheet.cell(row=row_num, column=3).value = value['date']
        sheet.cell(row=row_num, column=3).alignment = Alignment(vertical='top')
        sheet.cell(row=row_num, column=4).value = value['desc']
        sheet.cell(row=row_num, column=5).value = value['link']
        sheet.cell(row=row_num, column=5).alignment = Alignment(vertical='top')
        row_num += 1

    wb.save('Results %s.xlsx' % datetime.datetime.today().strftime('%Y-%m-%d'))
    print("Results saved in spreadsheet and spreadsheet closed.")

    return True


if __name__ == "__main__":

    # Start time to calculate and print elapsed time at the end
    start_run_time = time.time()

    # Make sure that our working directory is the directory
    # our .py file is located in:
    absolute_path = os.path.abspath(__file__)
    directory_name = os.path.dirname(absolute_path)
    os.chdir(directory_name)

    try:
        with open('runlog.txt', 'r') as old_time_log:
            content = old_time_log.readlines()
        last_line = content[-1].rstrip()

        print("Script last ran at: " + last_line)

        old_time = datetime.datetime.strptime(last_line, '%Y-%m-%d %H:%M:%S')

    except FileNotFoundError:
        print("runlog.txt not found.\nHow many days back should I start scraping?")
        start_time_input = int(input())
        start_time = datetime.datetime.strptime(datetime.datetime.now()
                                                .strftime("%Y-%m-%d %H:%M:%S"),
                                                '%Y-%m-%d %H:%M:%S')
        old_time = (start_time - datetime.timedelta(days=start_time_input))

    current_time = datetime.datetime.now()
    with open('runlog.txt', 'a') as new_time_log:
        new_time_log.write(current_time.strftime("%Y-%m-%d %H:%M:%S") + "\n")

    # Define arguments and call scraping functions.
    bsj_categories = ['operations', 'engineering']  # , 'design-ux',
    #'internships', 'sales', 'finance', 'product-management',
    #'contracting-positions', 'seeking-co-founders',
    #'hr-recruiting', 'marketing', 'other']

    for category in bsj_categories:
        bsj_results = scrape_bsj(category, old_time)
        if bsj_results:
            write_to_XLS(bsj_results)

    # This is a test string taken from the URL based on a number
    # of selected search criteria.
    stepstone_search_string = ("5/ergebnisliste.html"
                               "?ws=Berlin"
                               "&fu=6000000%2C1000000%2C7002000%2C7006000"
                               "&li=10&of=0"  # no. of search results to display
                               "&fci=419239&an=facets&fu=7008000"
                               "&fid=7008000&fn=categories&fa=select")
    stepstone_results = scrape_stepstone(stepstone_search_string, old_time)
    if stepstone_results:
        clean_results(stepstone_results)
        write_to_XLS(stepstone_results)

    # Print elapsed time.
    elapsed_run_time = time.time() - start_run_time
    print("All done.\nTime elapsed: " + time.strftime("%H:%M:%S",
                                                      time
                                                      .gmtime(elapsed_run_time)
                                                      ))
